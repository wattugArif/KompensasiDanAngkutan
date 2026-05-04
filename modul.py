import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Union
from collections import defaultdict

class DataFilterAndSelect:
    COLUMNS = ["Kode Testpit", "Grid", "Prospek", "Tanggal Sampling", "Total Koli", "Pemilik Lahan", "Pengangkut"]
           
    def __init__(self, source):
        if isinstance(source, pd.DataFrame):
            self.df = source
        elif isinstance(source, str): 
            try:
                self.df = pd.read_csv(source, encoding='utf-8')
            except Exception as e:
                raise ValueError(f"Failed to load CSV file '{source}': {e}")
        else:
            raise ValueError("Unsupported input type for DataFilterAndSelect")

        if "Tanggal Sampling" in self.df.columns:
            self.df["Tanggal Sampling"] = pd.to_datetime(
                self.df["Tanggal Sampling"], errors='coerce', dayfirst=True
            )

        self.cleanData = None
        
    def filter_and_select(self):
        # Menggunakan Kode Testpit untuk filter notna()
        filtered_df = self.df.loc[self.df["Kode Testpit"].notna(), self.COLUMNS].copy()
        if filtered_df.empty:
            raise ValueError("Kolom Contoh Error atau Kosong")
        
        self.cleanData = filtered_df
        return self.cleanData

class ConfigurationInput:
    def __init__(self):
        self.stage1 = None
        self.merged_data = None

    @staticmethod
    def _merge_stage_data(existing_df, new_df, subset):
        if existing_df is None:
            return new_df
        return pd.concat([existing_df, new_df], ignore_index=True).drop_duplicates(subset=subset)

    def process_stage1(self, cleanData):
        unique_locations = cleanData["Prospek"].unique()
        new_data = pd.DataFrame({
            "Lokasi": unique_locations,
            "Tanggal Mulai (2025-05-23)": pd.NaT,
            "Tanggal Selesai (2025-05-23)": pd.NaT,
            "Tanggal Gajian (2025-05-23)": pd.NaT,
            "Sistem Angkutan (Koli/Kilo)": np.nan,
        })

        self.stage1 = self._merge_stage_data(self.stage1, new_data, subset=["Lokasi"])
        return self.stage1

    def _filter_by_location_and_date(self, cleanData, stage1_data):
        filtered_chunks = []

        for _, row in stage1_data.iterrows():
            lokasi = row["Lokasi"]
            tgl_mulai = row["Tanggal Mulai (2025-05-23)"]
            tgl_selesai = row["Tanggal Selesai (2025-05-23)"]

            if pd.isna(tgl_mulai) and pd.isna(tgl_selesai):
                continue 

            filtered = cleanData[cleanData["Prospek"] == lokasi]

            if pd.notna(tgl_mulai) and pd.notna(tgl_selesai):
                filtered = filtered[(filtered["Tanggal Sampling"] >= tgl_mulai) & (filtered["Tanggal Sampling"] <= tgl_selesai)]
            elif pd.notna(tgl_mulai):
                filtered = filtered[filtered["Tanggal Sampling"] >= tgl_mulai]
            elif pd.notna(tgl_selesai):
                filtered = filtered[filtered["Tanggal Sampling"] <= tgl_selesai]

            if not filtered.empty:
                filtered_chunks.append(filtered)

        if filtered_chunks:
            result = pd.concat(filtered_chunks, ignore_index=True)
            return result
        else:
            return pd.DataFrame(columns=cleanData.columns)

    def process_stage(self, cleanData, stage1_data):
        result = self._filter_by_location_and_date(cleanData, stage1_data)
    
        # Create mapping from Lokasi to Sistem Angkutan
        mapping = stage1_data.set_index("Lokasi")["Sistem Angkutan (Koli/Kilo)"].to_dict()
    
        # Map to result DataFrame
        result["SistemAngkutan"] = result["Prospek"].map(mapping)
    
        self.merged_data = result
        return result

class PaymentCount:
    def __init__(self):
        self.df = None

    def set_data(self, df):
        self.df = df.copy()
        return self

    def harga_kompensasi(self, tarif=100000):
        # Menggunakan parameter tarif yang dilempar dari app.py
        self.df["Tarif Kompensasi"] = tarif
        return self

    def harga_angkutan(self):
        def hitung_angkutan(row):
            sistem = str(row.get("SistemAngkutan", "")).strip().lower()
            
            if sistem == "koli":
                return row.get("Total Koli", 0) * row.get("Pengangkut", 0) * 1000  
            elif sistem == "kilo":
                return row.get("Pengangkut", 0) * 1000   
            else:
                return 0  
    
        self.df["Tarif Angkutan"] = self.df.apply(hitung_angkutan, axis=1)
        return self

    def get_result(self):
        if "Kode Testpit" in self.df.columns:
            self.df = self.df.drop_duplicates(subset=["Kode Testpit"])
        return self.df 
    
    def get_pivot_summary(self):
        pivot_df = self.df.pivot_table(
            index=['Tanggal Sampling', 'Kode Testpit', 'Grid', 'Prospek', 'Pemilik Lahan'],
            values=['Tarif Kompensasi', 'Tarif Angkutan'],
            aggfunc='sum',
            fill_value=0
        ).reset_index()
        
        numeric_cols = list(pivot_df.select_dtypes(include='number').columns)
        pivot_df['Total'] = pivot_df[numeric_cols].sum(axis=1)
    
        total_values = pivot_df[numeric_cols + ['Total']].sum(axis=0)
        total_row = {col: total_values.get(col, None) for col in pivot_df.columns}
    
        for col in pivot_df.columns:
            if col not in numeric_cols and col != 'Total':
                total_row[col] = 'Total'
                
        for col in pivot_df.columns:
            if col not in numeric_cols + ['Total']:
                pivot_df[col] = pivot_df[col].astype(str)
    
        pivot_df = pd.concat([pivot_df, pd.DataFrame([total_row])], ignore_index=True)
        return pivot_df

class MultiPaymentExcel:
    def __init__(
        self,
        ws,
        data_rows: List[List[List[Union[str, float, int]]]],
        group_names: List[str],
        date_text: str = "Meliau, 11 Februari 2026",
        signers: dict = None ,
        receiver_title: str = "Area",
        mode: str = "kompensasi"
    ):
        if signers is None:
            signers = {
                "B": ("Dodi Prasetyo", "Keu. / Umum"),
                "D": ("Prya Arif Rahman", "Geologist"),
            }        
        self.ws = ws
        self.data_rows = data_rows
        self.group_names = group_names
        self.signers = signers
        self.date_text = date_text
        self.receiver_title = receiver_title
        self.mode = mode.lower()

    def generate_excel(self):
        ws = self.ws
        current_row = 1

        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        rp_style = NamedStyle(name="rupiah_style")
        rp_style.number_format = '"Rp."#,##0'
        rp_style.alignment = center_align
        rp_style.border = thin_border
        if "rupiah_style" not in ws.parent.named_styles:
            ws.parent.add_named_style(rp_style)

        # Konfigurasi per mode
        mode_config = {
            "kompensasi": {
                "title": "PEMBAYARAN KOMPENSASI LAHAN",
                "uraian": "Untuk Pembayaran Kompensasi Lahan sbb :",
                "headers": ["No", "Tgl. Selesai", "Kode Tespit", "Grid", "Pemilik Lahan", "Harga Kompensasi", "Total Kompensasi", "TTD"],
                "harga_col": 7,
                "subtotal": True,
            },
            "angkutan": {
                "title": "PEMBAYARAN ANGKUTAN SAMPEL",
                "uraian": "Untuk Pembayaran Angkutan Sampel sbb :",
                "headers": ["No", "Tgl. Selesai", "Kode Tespit", "Grid", "Pemilik Lahan", "Harga Angkutan", "TTD"],
                "harga_col": 7,
                "subtotal": False,
            }
        }

        config = mode_config[self.mode]

        for table_index, table_rows in enumerate(self.data_rows):
            group_name = self.group_names[table_index]

            for title in ["BUKTI PEMBAYARAN", config["title"]]:
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=len(config["headers"]) + 1)
                cell = ws.cell(row=current_row, column=2, value=title)
                cell.font = Font(bold=True, size=14 if title == "BUKTI PEMBAYARAN" else 12)
                cell.alignment = center_align
                current_row += 1

            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=len(config["headers"]) + 1)
            ws.cell(row=current_row, column=2, value="Sudah Terima Dari : Tim Eksplorasi Bauksit Kalbar").alignment = left_align
            current_row += 2

            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=len(config["headers"]) + 1)
            ws.cell(row=current_row, column=2, value=config["uraian"]).alignment = left_align
            current_row += 1

            for col_index, header in enumerate(config["headers"], start=2):
                cell = ws.cell(row=current_row, column=col_index, value=header)
                cell.font = bold_font
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = header_fill
                ws.column_dimensions[get_column_letter(col_index)].width = len(header) + 5
            current_row += 1

            total_harga = 0
            start_data_row = current_row
            
            # Sort data berdasarkan Pemilik Lahan (case-insensitive)
            table_rows.sort(key=lambda x: str(x[4]).strip().lower())

            for i, row_data in enumerate(table_rows):
                row_data[0] = i + 1
                while len(row_data) < len(config["headers"]):
                    row_data.append("")
                for j, val in enumerate(row_data):
                    col = j + 2
                    cell = ws.cell(row=current_row, column=col, value=val)
                    cell.alignment = center_align
                    cell.border = thin_border
                    if col == config["harga_col"]:
                        cell.style = "rupiah_style"
                if isinstance(row_data[config["harga_col"] - 2], (int, float)):
                    total_harga += row_data[config["harga_col"] - 2]
                current_row += 1

            # Subtotal and merging logic for kompensasi mode
            if self.mode == "kompensasi":
                # Create a map to store all rows belonging to the same owner
                owner_row_map = defaultdict(list)
                for idx, row_data in enumerate(table_rows):
                    # Index 4 is "Pemilik Lahan" column
                    # Gunakan .lower() agar grouping sama persis dengan urutan sorting
                    owner = str(row_data[4]).strip().lower()
                    owner_row_map[owner].append(start_data_row + idx)
                
                for owner, rows in owner_row_map.items():
                    subtotal = 0
                    for row in rows:
                        # Column 7 is Excel Col G ("Harga Kompensasi")
                        cell = ws.cell(row=row, column=7)
                        if isinstance(cell.value, (int, float)):
                            subtotal += cell.value
                    
                    # First row of range to merge
                    first_row = rows[0]
                    # Last row of range to merge
                    last_row = rows[-1]

                    # Write total subtotal
                    ws.cell(row=first_row, column=8, value=int(subtotal)).style = "rupiah_style"
                    
                    # Perform merging for "Total Kompensasi" & "TTD"
                    if len(rows) > 1:
                        ws.merge_cells(start_row=first_row, start_column=8, end_row=last_row, end_column=8)
                        ws.merge_cells(start_row=first_row, start_column=9, end_row=last_row, end_column=9)
                    
                    # Apply center alignment
                    ws.cell(row=first_row, column=8).alignment = center_align
                    ws.cell(row=first_row, column=9).alignment = center_align

            total_label_col = config["harga_col"] - 1
            label_cell = ws.cell(row=current_row, column=total_label_col, value="TOTAL")
            label_cell.font = bold_font
            label_cell.alignment = center_align
            label_cell.border = thin_border

            total_cell = ws.cell(row=current_row, column=config["harga_col"], value=total_harga)
            total_cell.style = "rupiah_style"

            current_row += 2

            if self.mode in ["kompensasi", "angkutan"]:
                ws.cell(row=current_row, column=8, value=self.date_text).alignment = left_align
                current_row += 1
                ws.cell(row=current_row, column=2, value="Dibayar Oleh,").alignment = center_align
                ws.cell(row=current_row, column=5, value="Pet. Lapangan,").alignment = center_align
                ws.cell(row=current_row, column=8, value="Lokasi,").alignment = center_align
                current_row += 5
            
                ws.cell(row=current_row, column=2, value=self.signers["B"][0]).alignment = center_align
                ws.cell(row=current_row, column=5, value=self.signers["D"][0]).alignment = center_align
                ws.cell(row=current_row, column=8, value=group_name).alignment = center_align
                current_row += 1
            
                ws.cell(row=current_row, column=2, value=self.signers["B"][1]).alignment = center_align
                ws.cell(row=current_row, column=5, value=self.signers["D"][1]).alignment = center_align
                ws.cell(row=current_row, column=8, value=self.receiver_title).alignment = center_align
                current_row += 4
                
class PaymentExcelBuilder:
    def __init__(self, df: pd.DataFrame):
        self.df = df.sort_values(by=['Prospek']).copy()
        self.df.fillna(0, inplace=True)
        
    def _group_data(self, group_col, columns, rename_map, values_structure, mode: str):
        raw_data = self.df[columns].rename(columns=rename_map)
        raw_data_list = raw_data.to_dict(orient="records")
    
        grouped = defaultdict(list)
        for entry in raw_data_list:
            row = [
                None,
                *[entry[col] for col in values_structure]
            ]
    
            if mode in ("kompensasi", "angkutan"):
                row.append("")
    
            grouped[entry[group_col]].append(row)
    
        tables = list(grouped.values())
        names = list(grouped.keys())
        return tables, names
        
    def create_multi_payment_excel(
        self,
        output_file: str,
        date_text: str = "Meliau, 11 Februari 2026",
        signers: dict = None
    ):
        if signers is None:
            signers = {
                "B": ("Dodi Prasetyo", "Keu. / Umum"),
                "D": ("Prya Arif Rahman", "Geologist"),
            }
        wb = Workbook()
        wb.remove(wb.active)

        configs = [
            {
                "sheet": "Kompensasi",
                "mode": "kompensasi",
                "group_col": "Prospek",
                "columns": ["Prospek", "Tanggal Sampling", "Kode Testpit", "Grid", "Pemilik Lahan", "Tarif Kompensasi"],
                "rename": {"Tarif Kompensasi": "harga"},
                "values": ["Tanggal Sampling", "Kode Testpit", "Grid", "Pemilik Lahan", "harga"]
            },
            {
                "sheet": "Angkutan",
                "mode": "angkutan",
                "group_col": "Prospek",
                "columns": ["Prospek", "Tanggal Sampling", "Kode Testpit", "Grid", "Pemilik Lahan", "Tarif Angkutan"],
                "rename": {"Tarif Angkutan": "harga"},
                "values": ["Tanggal Sampling", "Kode Testpit", "Grid", "Pemilik Lahan", "harga"]
            }
        ]

        for config in configs:
            tables, names = self._group_data(
                config["group_col"],
                config["columns"],
                config["rename"],
                config["values"],
                config["mode"]
            )
            ws = wb.create_sheet(config["sheet"])
            report = MultiPaymentExcel(
                ws,
                tables,
                names,
                date_text=date_text,
                signers=signers,
                mode=config["mode"]
            )
            report.generate_excel()

        wb.save(output_file)
